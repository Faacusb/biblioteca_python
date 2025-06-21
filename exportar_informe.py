import json
import os
from openpyxl import Workbook
from libros import archivo_libros
from usuarios import archivo_usuarios

archivo_prestamos = 'prestamos.json'

def exportar_informe():
    wb = Workbook()

    # ----------- Hoja 1: Libros -----------
    if os.path.exists(archivo_libros):
        with open(archivo_libros, 'r', encoding='utf-8') as f:
            libros = json.load(f)
    else:
        libros = []

    hoja_libros = wb.active
    hoja_libros.title = "Libros"
    hoja_libros.append(["Título", "Autor", "Año", "Prestado", "Socio Préstamo"])

    for libro in libros:
        hoja_libros.append([
            libro.get("titulo", ""),
            libro.get("autor", ""),
            libro.get("año", ""),
            "Sí" if libro.get("prestado") else "No",
            libro.get("socio_prestamo", "")
        ])

    # ----------- Hoja 2: Usuarios -----------
    if os.path.exists(archivo_usuarios):
        with open(archivo_usuarios, 'r', encoding='utf-8') as f:
            usuarios = json.load(f)
    else:
        usuarios = []

    hoja_usuarios = wb.create_sheet(title="Usuarios")
    hoja_usuarios.append(["Socio", "Nombre", "Apellido", "Tipo Doc", "Documento", "Dirección"])

    for usuario in usuarios:
        hoja_usuarios.append([
            usuario.get("socio", ""),
            usuario.get("nombre", ""),
            usuario.get("apellido", ""),
            usuario.get("tipo_doc", ""),
            usuario.get("documento", ""),
            usuario.get("direccion", "")
        ])

    # ----------- Hoja 3: Préstamos -----------
    if os.path.exists(archivo_prestamos):
        with open(archivo_prestamos, 'r', encoding='utf-8') as f:
            prestamos = json.load(f)
    else:
        prestamos = []

    hoja_prestamos = wb.create_sheet(title="Préstamos")
    hoja_prestamos.append(["Socio", "Nombre", "Título del Libro", "Fecha Préstamo", "Estado", "Fecha Devolución"])

    for prestamo in prestamos:
        hoja_prestamos.append([
            prestamo.get("socio", ""),
            prestamo.get("nombre", ""),
            prestamo.get("titulo_libro", ""),
            prestamo.get("fecha_prestamo", ""),
            prestamo.get("estado", ""),
            prestamo.get("fecha_devolucion", "") if prestamo.get("estado") == "devuelto" else ""
        ])

    # ----------- Guardar archivo Excel -----------
    archivo_salida = "informe.xlsx"
    wb.save(archivo_salida)
    print(f"Informe exportado con éxito a '{archivo_salida}'.")

if __name__ == "__main__":
    exportar_informe()
